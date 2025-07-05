from dotenv import load_dotenv
import os
from flask import Flask, redirect, url_for, render_template, session, send_file, flash, request, jsonify, Response, make_response
from auth import auth_bp 
from models import db, Usuario
from form import ContactoForm
import pandas as pd
from flask_login import LoginManager, login_required, current_user, AnonymousUserMixin
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter
from email_utils import mail, init_serializer
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash
import logging
from datetime import timedelta, datetime
import pytz
from flask_wtf.csrf import CSRFProtect
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
import glob

# Configurar logging primero
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Cargar variables de entorno
logging.info("Intentando cargar archivo .env...")
load_dotenv(override=True)  # Forzar recarga de variables
logging.info("Archivo .env cargado")

# Verificar variables críticas
logging.info("Variables de entorno cargadas:")
logging.info(f"MAIL_USERNAME: {os.environ.get('MAIL_USERNAME')}")
logging.info(f"MAIL_PASSWORD: {'Configurado' if os.environ.get('MAIL_PASSWORD') else 'NO CONFIGURADO'}")
logging.info(f"MAIL_DEFAULT_SENDER: {os.environ.get('MAIL_DEFAULT_SENDER')}")

class Anonymous(AnonymousUserMixin):
    @property
    def is_authenticated(self):
        return False
    
    @property
    def is_active(self):
        return False
    
    @property
    def is_anonymous(self):
        return True
    
    def get_id(self):
        return None

# Inicializar Flask y sus extensiones
gestor = Flask(__name__)
gestor.config.from_object('config.Config')

# Configuración de sesión
gestor.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)
gestor.config['SESSION_COOKIE_SECURE'] = False  # Cambiado a False para desarrollo local
gestor.config['SESSION_COOKIE_HTTPONLY'] = True
gestor.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
gestor.config['SESSION_PROTECTION'] = 'basic'  # Cambiado a 'basic' en lugar de 'strong'
gestor.config['REMEMBER_COOKIE_DURATION'] = timedelta(days=14)
gestor.config['REMEMBER_COOKIE_REFRESH_EACH_REQUEST'] = True

# Inicializar extensiones
csrf = CSRFProtect(gestor)
db.init_app(gestor)
migrate = Migrate(gestor, db)
mail.init_app(gestor)
init_serializer(gestor)

# Registrar blueprints
gestor.register_blueprint(auth_bp)

# Configurar Flask-Login
login_manager = LoginManager()
login_manager.init_app(gestor)
login_manager.login_view = "auth.login"
login_manager.login_message = "Por favor inicia sesión para acceder a esta página."
login_manager.login_message_category = "info"
login_manager.login_message_category = "info"
login_manager.session_protection = "basic"  # Cambiado a 'basic'
login_manager.refresh_view = "auth.login"
login_manager.needs_refresh_message = "Por favor inicia sesión nuevamente para continuar."
login_manager.needs_refresh_message_category = "info"

@login_manager.user_loader
def load_user(user_id):
    if user_id is None:
        return None
    try:
        user = db.session.get(Usuario, int(user_id))
        if user:
            # Renovar la sesión
            session.permanent = True
            session['user_id'] = user.id
            session.modified = True
        return user
    except (ValueError, TypeError):
        return None

# Crear la base de datos SQLite
with gestor.app_context():
    from models import Contacto
    try:
        # Asegurar que el directorio instance existe para SQLite
        instance_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance')
        if not os.path.exists(instance_path):
            os.makedirs(instance_path)
            logging.info(f"Directorio instance creado: {instance_path}")
        
        db.create_all()
        logging.info("Base de datos SQLite creada/verificada exitosamente")
        
    except Exception as e:
        logging.error(f"Error al crear la base de datos SQLite: {str(e)}")
        # Continuar sin la base de datos para evitar que la aplicación falle completamente

# Configurar manejo de usuarios anónimos
login_manager.anonymous_user = Anonymous

def limpiar_archivos_temporales():
    """Eliminar archivos Excel temporales antiguos (más de 1 hora)"""
    try:
        import os
        import time
        
        # Limpiar archivos .xlsx en el directorio raíz
        for archivo in glob.glob("*.xlsx"):
            if os.path.exists(archivo):
                tiempo_modificacion = os.path.getmtime(archivo)
                tiempo_actual = time.time()
                # Eliminar archivos más antiguos de 1 hora
                if tiempo_actual - tiempo_modificacion > 3600:
                    os.remove(archivo)
                    logging.info(f"Archivo temporal eliminado: {archivo}")
        
        # Limpiar archivos en el directorio temp
        temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
        if os.path.exists(temp_dir):
            for archivo in glob.glob(os.path.join(temp_dir, "*.xlsx")):
                if os.path.exists(archivo):
                    tiempo_modificacion = os.path.getmtime(archivo)
                    tiempo_actual = time.time()
                    # Eliminar archivos más antiguos de 1 hora
                    if tiempo_actual - tiempo_modificacion > 3600:
                        os.remove(archivo)
                        logging.info(f"Archivo temporal eliminado: {archivo}")
    except Exception as e:
        logging.error(f"Error al limpiar archivos temporales: {str(e)}")

@gestor.route('/admin')
@login_required
def admin():
    if not current_user.is_admin:
        return redirect(url_for('auth.login'))
    contactos = Contacto.query.all()
    return render_template('admin.html', contactos=contactos, now=datetime.now)

@gestor.route('/exportar_contactos')
@login_required
def exportar_contactos():
    if not current_user.is_admin:
        return redirect(url_for('auth.login'))
    
    # Limpiar archivos temporales antes de generar nuevo archivo
    limpiar_archivos_temporales()
    
    # Forzar recarga de datos desde la base de datos
    db.session.expire_all()
    contactos = Contacto.query.all()
    
    # Log para verificar cuántos contactos se están exportando
    logging.info(f"Exportando {len(contactos)} contactos para admin {current_user.email}")
    
    # Agregar timestamp al nombre del archivo para evitar caché
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"contactos_{timestamp}.xlsx"
    
    # Función para normalizar texto
    def normalize_text(text):
        if text:
            # Convertir a minúsculas y capitalizar la primera letra
            return text.strip().lower().capitalize()
        return text
    
    data = [{
        "Origen": normalize_text(c.origen),
        "Cobertura Actual": normalize_text(c.cobertura_actual_otra) if c.cobertura_actual == 'otros' else c.cobertura_actual,
        "¿Por qué no toma la cobertura?": normalize_text(c.promocion) if c.promocion else "No especificado",
        "Privado/Desregulado": c.privadoDesregulado,
        "Apellido y nombre": c.apellido_nombre,
        "Correo electrónico": c.correo_electronico,
        "Edad titular": c.edad_titular,
        "Teléfono": c.telefono,
        "Grupo familiar": c.grupo_familiar,
        "Plan ofrecido": c.plan_ofrecido,
        "Estado": c.estado,
        "Observaciones": c.observaciones,
        "Cónyuge": c.conyuge,
        "Edad cónyuge": c.conyuge_edad,
        "Fecha de carga": c.created_at.strftime("%d/%m/%Y"),
        "Usuario cargador": c.usuario.email
    } for c in contactos]
    
    df = pd.DataFrame(data)
    
    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    
    # Configurar estilos generales
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    month_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    month_font = Font(bold=True, size=11, color="1F4E78")
    
    # Crear la hoja de datos
    ws_data = wb.active
    ws_data.title = "Base de Datos"

    # Ordenar los datos por fecha de carga
    df['Mes de carga'] = pd.to_datetime(df['Fecha de carga'], format='%d/%m/%Y').dt.strftime('%B %Y')
    df = df.sort_values('Fecha de carga')

    # Agrupar por mes
    grupos_mes = df.groupby('Mes de carga')

    # Fila actual para ir agregando datos
    current_row = 1

    for mes, grupo in grupos_mes:
        # Agregar encabezado del mes
        ws_data.merge_cells(f'A{current_row}:P{current_row}')
        mes_cell = ws_data[f'A{current_row}']
        mes_cell.value = f"CONTACTOS CARGADOS EN {mes.upper()}"
        mes_cell.fill = month_fill
        mes_cell.font = month_font
        mes_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        # Agregar encabezados de columnas
        headers = [
            "Origen", "Cobertura Actual", "¿Por qué no toma la cobertura?", "Privado/Desregulado",
            "Apellido y nombre", "Correo electrónico", "Edad titular",
            "Teléfono", "Grupo familiar", "Plan ofrecido",
            "Estado", "Observaciones", "Cónyuge", "Edad cónyuge", "Fecha de carga", "Usuario cargador"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        current_row += 1

        # Agregar datos del grupo
        for _, row in grupo.iterrows():
            ws_data.cell(row=current_row, column=1).value = row['Origen']
            ws_data.cell(row=current_row, column=2).value = row['Cobertura Actual']
            ws_data.cell(row=current_row, column=3).value = row['¿Por qué no toma la cobertura?']
            ws_data.cell(row=current_row, column=4).value = row['Privado/Desregulado']
            ws_data.cell(row=current_row, column=5).value = row['Apellido y nombre']
            ws_data.cell(row=current_row, column=6).value = row['Correo electrónico']
            ws_data.cell(row=current_row, column=7).value = row['Edad titular']
            ws_data.cell(row=current_row, column=8).value = row['Teléfono']
            ws_data.cell(row=current_row, column=9).value = row['Grupo familiar']
            ws_data.cell(row=current_row, column=10).value = row['Plan ofrecido']
            ws_data.cell(row=current_row, column=11).value = row['Estado']
            ws_data.cell(row=current_row, column=12).value = row['Observaciones']
            ws_data.cell(row=current_row, column=13).value = row['Cónyuge']
            ws_data.cell(row=current_row, column=14).value = row['Edad cónyuge']
            ws_data.cell(row=current_row, column=15).value = row['Fecha de carga']
            ws_data.cell(row=current_row, column=16).value = row['Usuario cargador']

            # Aplicar estilo según el estado
            estado_cell = ws_data.cell(row=current_row, column=11)
            if row['Estado'] == 'abierto':
                estado_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                estado_cell.font = Font(color="006100")
            elif row['Estado'] == 'cerrado':
                estado_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                estado_cell.font = Font(color="9C0006")
            elif row['Estado'] == 'vendido':
                estado_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                estado_cell.font = Font(color="1F4E78")
            elif row['Estado'] == 'vendido':
                estado_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                estado_cell.font = Font(color="1F4E78")

            current_row += 1

        # Agregar espacio entre grupos
        current_row += 1

    # Ajustar ancho de columnas
    for column in ws_data.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_data.column_dimensions[column_letter].width = min(adjusted_width, 30)

    # Agregar bordes a todas las celdas con datos
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws_data.iter_rows(min_row=1, max_row=ws_data.max_row):
        for cell in row:
            if cell.value is not None:
                cell.border = border
                if not cell.alignment:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Agregar una nueva hoja de resumen por mes
    ws_monthly = wb.create_sheet(title="Resumen Mensual")
    
    # Título
    ws_monthly.merge_cells('A1:F1')
    ws_monthly['A1'] = "RESUMEN DE CONTACTOS POR MES"
    ws_monthly['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws_monthly['A1'].alignment = Alignment(horizontal="center")
    
    # Encabezados
    headers = ["Mes", "Total Contactos", "Planes Vendidos", "Abiertos", "Cerrados", "Tasa de Efectividad"]
    for col, header in enumerate(headers, 1):
        cell = ws_monthly.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Datos por mes
    current_row = 4
    for mes, grupo in grupos_mes:
        total = len(grupo)
        vendidos = len(grupo[grupo['Estado'] == 'vendido'])
        abiertos = len(grupo[grupo['Estado'] == 'abierto'])
        cerrados = len(grupo[grupo['Estado'] == 'cerrado'])
        tasa = (vendidos / total * 100) if total > 0 else 0
        
        ws_monthly[f'A{current_row}'] = mes
        ws_monthly[f'B{current_row}'] = total
        ws_monthly[f'C{current_row}'] = vendidos
        ws_monthly[f'D{current_row}'] = abiertos
        ws_monthly[f'E{current_row}'] = cerrados
        ws_monthly[f'F{current_row}'] = f"{tasa:.1f}%"
        
        # Formato
        for col in range(1, 7):
            cell = ws_monthly.cell(row=current_row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        current_row += 1
    
    # Ajustar ancho de columnas
    for column in ws_monthly.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_monthly.column_dimensions[column_letter].width = min(adjusted_width, 30)

    # Agregar hoja de estadísticas de campos
    ws_stats = wb.create_sheet(title="Estadísticas de Campos")
    
    # Título
    ws_stats.merge_cells('A1:D1')
    ws_stats['A1'] = "DISTRIBUCIÓN DE OPCIONES SELECCIONADAS"
    ws_stats['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws_stats['A1'].alignment = Alignment(horizontal="center")
    
    # Encabezados
    headers = ["Campo", "Opción", "Cantidad", "Porcentaje"]
    for col, header in enumerate(headers, 1):
        cell = ws_stats.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Campos a analizar
    campos = {
        'Origen': 'Origen',
        'Cobertura Actual': 'Cobertura Actual',
        '¿Por qué no toma la cobertura?': '¿Por qué no toma la cobertura?',
        'Privado/Desregulado': 'Privado/Desregulado',
        'Estado': 'Estado'
    }
    
    current_row = 4
    
    # Analizar cada campo
    for campo_df, nombre_campo in campos.items():
        # Agregar título del campo
        ws_stats.merge_cells(f'A{current_row}:D{current_row}')
        campo_cell = ws_stats[f'A{current_row}']
        campo_cell.value = f"DISTRIBUCIÓN DE {nombre_campo.upper()}"
        campo_cell.fill = month_fill
        campo_cell.font = month_font
        campo_cell.alignment = Alignment(horizontal="center")
        current_row += 1
        
        # Calcular distribución
        total = len(df)
        distribucion = df[campo_df].value_counts()
        
        # Agregar datos
        for opcion, cantidad in distribucion.items():
            porcentaje = (cantidad / total * 100)
            
            ws_stats[f'A{current_row}'] = nombre_campo
            ws_stats[f'B{current_row}'] = opcion
            ws_stats[f'C{current_row}'] = cantidad
            ws_stats[f'D{current_row}'] = f"{porcentaje:.1f}%"
            
            # Formato
            for col in range(1, 5):
                cell = ws_stats.cell(row=current_row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            
            current_row += 1
        
        # Espacio entre campos
        current_row += 1
    
    # Ajustar ancho de columnas
    for column in ws_stats.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_stats.column_dimensions[column_letter].width = min(adjusted_width, 30)

    # Agregar hoja de ranking de usuarios
    ws_ranking = wb.create_sheet(title="Ranking de Usuarios")
    
    # Título
    ws_ranking.merge_cells('A1:E1')
    ws_ranking['A1'] = "RANKING DE USUARIOS POR VENTAS"
    ws_ranking['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws_ranking['A1'].alignment = Alignment(horizontal="center")
    
    # Encabezados
    headers = ["Posición", "Usuario", "Contactos Vendidos", "Total Contactos", "Tasa de Éxito"]
    for col, header in enumerate(headers, 1):
        cell = ws_ranking.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Calcular ranking de usuarios
    ranking_usuarios = df.groupby('Usuario cargador').agg({
        'Apellido y nombre': 'count',  # Total de contactos
        'Estado': lambda x: (x == 'vendido').sum()  # Contactos vendidos
    }).rename(columns={
        'Apellido y nombre': 'Total Contactos',
        'Estado': 'Contactos Vendidos'
    })
    
    # Calcular tasa de éxito
    ranking_usuarios['Tasa de Éxito'] = (ranking_usuarios['Contactos Vendidos'] / ranking_usuarios['Total Contactos'] * 100).round(1)
    
    # Ordenar por contactos vendidos (de mayor a menor) - Los mejores vendedores primero
    ranking_usuarios = ranking_usuarios.sort_values('Contactos Vendidos', ascending=False)
    
    # Agregar datos al Excel
    current_row = 4
    for posicion, (usuario, datos) in enumerate(ranking_usuarios.iterrows(), 1):
        ws_ranking[f'A{current_row}'] = posicion
        ws_ranking[f'B{current_row}'] = usuario
        ws_ranking[f'C{current_row}'] = datos['Contactos Vendidos']
        ws_ranking[f'D{current_row}'] = datos['Total Contactos']
        ws_ranking[f'E{current_row}'] = f"{datos['Tasa de Éxito']:.1f}%"
        
        # Formato especial para el top 3
        if posicion <= 3:
            # Fondo dorado para el 1er lugar
            if posicion == 1:
                fill_color = "FFD700"  # Dorado
                font_color = "000000"
            # Fondo plateado para el 2do lugar
            elif posicion == 2:
                fill_color = "C0C0C0"  # Plateado
                font_color = "000000"
            # Fondo bronce para el 3er lugar
            else:
                fill_color = "CD7F32"  # Bronce
                font_color = "FFFFFF"
            
            for col in range(1, 6):
                cell = ws_ranking.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = Font(bold=True, color=font_color)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
        else:
            # Formato normal para el resto
            for col in range(1, 6):
                cell = ws_ranking.cell(row=current_row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
        
        current_row += 1
    
    
    # Ajustar ancho de columnas
    for column in ws_ranking.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_ranking.column_dimensions[column_letter].width = min(adjusted_width, 30)

    # Guardar el archivo en un directorio temporal
    import os
    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
    file_path = os.path.join(temp_dir, filename)
    wb.save(file_path)

    logging.info(f"¿Existe el archivo? {os.path.exists(file_path)} - {file_path}")

    try:
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logging.error(f"Error al enviar archivo: {str(e)}")
        return jsonify({'error': 'Error al enviar el archivo'}), 500

@gestor.route('/usuario', methods=["GET", "POST"])
@login_required
def usuario():
    logging.info(f"Accediendo a ruta /usuario - Usuario autenticado: {current_user.is_authenticated}")
    
    # Verificar autenticación
    if not current_user.is_authenticated:
        logging.error("Usuario no autenticado en ruta protegida")
        flash("Por favor inicia sesión para continuar", "warning")
        return redirect(url_for('auth.login'))
    
    # Renovar la sesión
    session.permanent = True
    session['user_id'] = current_user.id
    session['_fresh'] = True
    session.modified = True
    
    form = ContactoForm()
    
    if request.method == "POST":
        logging.info("Recibiendo POST para agregar contacto")
        logging.info(f"Form data: {request.form}")
        
        # Verificar campos requeridos
        campos_requeridos = [
            ('origen', 'Origen'),
            ('privadoDesregulado', 'Privado/Desregulado'),
            ('apellido_nombre', 'Apellido y nombre'),
            ('correo_electronico', 'Correo electrónico'),
            ('edad_titular', 'Edad titular'),
            ('telefono', 'Teléfono'),
            ('grupo_familiar', 'Grupo familiar'),
            ('plan_ofrecido', 'Plan ofrecido'),
            ('estado', 'Estado')
        ]
        
        campos_vacios = []
        for campo, nombre in campos_requeridos:
            valor = getattr(form, campo).data
            if not valor or (isinstance(valor, str) and not valor.strip()):
                campos_vacios.append(nombre)
        
        if campos_vacios:
            flash(f"Por favor complete los siguientes campos obligatorios: {', '.join(campos_vacios)}", "danger")
            return render_template('usuario.html', form=form, contactos_usuario=contactos_usuario)
        
        if form.validate_on_submit():
            logging.info("Formulario validado correctamente")
            try:
                # Procesar campos con opción "otros"
                cobertura_final = form.cobertura_actual_otra.data if form.cobertura_actual.data == 'otros' else form.cobertura_actual.data
                
                # Procesar campo de cónyuge
                conyuge_edad_final = form.conyuge_edad.data if form.conyuge.data == 'con conyuge' else 'Sin cónyuge'
                
                nuevo_contacto = Contacto(
                    usuario_id=current_user.id,
                    origen=form.origen.data,
                    cobertura_actual=cobertura_final,
                    cobertura_actual_otra=form.cobertura_actual_otra.data if form.cobertura_actual.data == 'otros' else None,
                    promocion=form.promocion.data,
                    privadoDesregulado=form.privadoDesregulado.data,
                    apellido_nombre=form.apellido_nombre.data,
                    correo_electronico=form.correo_electronico.data,
                    edad_titular=form.edad_titular.data,
                    telefono=form.telefono.data,
                    grupo_familiar=form.grupo_familiar.data,
                    plan_ofrecido=form.plan_ofrecido.data,
                    fecha=datetime.now(pytz.timezone('America/Argentina/Mendoza')).strftime("%d/%m/%Y"),
                    estado=form.estado.data,
                    observaciones=form.observaciones.data,
                    conyuge=form.conyuge.data,
                    conyuge_edad=conyuge_edad_final
                )
                db.session.add(nuevo_contacto)
                db.session.commit()
                
                flash("Contacto agregado correctamente", "success")
                logging.info("Contacto agregado exitosamente")
                return redirect(url_for('usuario'))
                
            except Exception as e:
                logging.error(f"Error al agregar contacto: {str(e)}")
                db.session.rollback()
                flash("Error al agregar el contacto. Por favor intenta nuevamente.", "danger")
        else:
            logging.error(f"Errores en el formulario: {form.errors}")
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f"Error en {getattr(form, field).label.text}: {error}", "danger")

    try:
        contactos_usuario = Contacto.query.filter_by(usuario_id=current_user.id).all()
    except Exception as e:
        logging.error(f"Error al obtener contactos: {str(e)}")
        contactos_usuario = []
        flash("Error al cargar los contactos", "warning")

    return render_template('usuario.html', form=form, contactos_usuario=contactos_usuario)

@gestor.route('/actualizar_promocion', methods=['POST'])
@login_required
def actualizar_promocion():
    try:
        logging.info(f"Recibiendo solicitud para actualizar promoción - Usuario: {current_user.id}")
        contacto_id = request.form.get('contacto_id')
        nueva_promocion = request.form.get('nueva_promocion')
        
        logging.info(f"Datos recibidos - contacto_id: {contacto_id}, nueva_promocion: {nueva_promocion}")
        
        if not contacto_id:
            logging.warning("ID de contacto no proporcionado")
            return jsonify({'success': False, 'message': 'ID de contacto requerido'})
        
        # Verificar que el contacto pertenece al usuario actual
        contacto = Contacto.query.filter_by(id=contacto_id, usuario_id=current_user.id).first()
        
        if not contacto:
            logging.warning(f"Contacto no encontrado - ID: {contacto_id}, Usuario: {current_user.id}")
            return jsonify({'success': False, 'message': 'Contacto no encontrado'})
        
        # Verificar que el estado sea "cerrado"
        if contacto.estado != 'cerrado':
            logging.warning(f"Contacto no está en estado cerrado - Estado actual: {contacto.estado}")
            return jsonify({'success': False, 'message': 'Solo se puede actualizar la promoción en contactos cerrados'})
        
        # Validar opciones de promoción
        opciones_validas = [
            'promocion sancor', 'promocion medicus', 'promocion omint', 'promocion prevencion',
            'promocion smg', 'promocion medife', 'osde', 'servicios insatisfactorios',
            'no puede pagarlo', 'otros prepagos', ''
        ]
        
        if nueva_promocion not in opciones_validas:
            logging.warning(f"Promoción no válida: {nueva_promocion}")
            return jsonify({'success': False, 'message': 'Opción de promoción no válida'})
        
        # Actualizar la promoción
        contacto.promocion = nueva_promocion if nueva_promocion else None
        db.session.commit()
        
        logging.info(f"Promoción actualizada exitosamente - Nueva promoción: {nueva_promocion}")
        
        return jsonify({
            'success': True, 
            'message': 'Promoción actualizada correctamente',
            'nueva_promocion': nueva_promocion
        })
        
    except Exception as e:
        logging.error(f"Error al actualizar promoción del contacto: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Error al actualizar la promoción'})

@gestor.route('/cambiar_estado_contacto', methods=['POST'])
@login_required
def cambiar_estado_contacto():
    try:
        logging.info(f"Recibiendo solicitud para cambiar estado - Usuario: {current_user.id}")
        contacto_id = request.form.get('contacto_id')
        nuevo_estado = request.form.get('nuevo_estado')
        
        logging.info(f"Datos recibidos - contacto_id: {contacto_id}, nuevo_estado: {nuevo_estado}")
        
        if not contacto_id or not nuevo_estado:
            logging.warning("Datos incompletos en la solicitud")
            return jsonify({'success': False, 'message': 'Datos incompletos'})
        
        # Verificar que el contacto pertenece al usuario actual
        contacto = Contacto.query.filter_by(id=contacto_id, usuario_id=current_user.id).first()
        
        if not contacto:
            logging.warning(f"Contacto no encontrado - ID: {contacto_id}, Usuario: {current_user.id}")
            return jsonify({'success': False, 'message': 'Contacto no encontrado'})
        
        logging.info(f"Contacto encontrado - Estado actual: {contacto.estado}")
        
        # Validar que el estado sea válido
        estados_validos = ['abierto', 'cerrado', 'no responde', 'vendido']
        if nuevo_estado not in estados_validos:
            logging.warning(f"Estado no válido: {nuevo_estado}")
            return jsonify({'success': False, 'message': 'Estado no válido'})
        
        # Actualizar el estado
        contacto.estado = nuevo_estado
        db.session.commit()
        
        logging.info(f"Estado actualizado exitosamente - Nuevo estado: {nuevo_estado}")
        
        return jsonify({
            'success': True, 
            'message': f'Estado actualizado a {nuevo_estado.capitalize()}',
            'nuevo_estado': nuevo_estado
        })
        
    except Exception as e:
        logging.error(f"Error al cambiar estado del contacto: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Error al actualizar el estado'})

@gestor.route('/exportar_mis_contactos_simple')
@login_required
def exportar_mis_contactos_simple():
    """Versión simplificada de exportación para PythonAnywhere"""
    try:
        logging.info(f"Exportación simple para usuario {current_user.email}")
        
        # Obtener contactos
        contactos = Contacto.query.filter_by(usuario_id=current_user.id).all()
        
        # Crear archivo CSV simple como fallback
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Encabezados
        headers = [
            "Origen", "Cobertura Actual", "¿Por qué no toma la cobertura?", 
            "Privado/Desregulado", "Apellido y nombre", "Correo electrónico",
            "Edad titular", "Teléfono", "Grupo familiar", "Plan ofrecido",
            "Estado", "Observaciones", "Cónyuge", "Edad cónyuge", "Fecha de carga"
        ]
        writer.writerow(headers)
        
        # Datos
        for c in contactos:
            row = [
                c.origen or "",
                c.cobertura_actual_otra if c.cobertura_actual == 'otros' else (c.cobertura_actual or ""),
                c.promocion or "No especificado",
                c.privadoDesregulado or "",
                c.apellido_nombre or "",
                c.correo_electronico or "",
                c.edad_titular or "",
                c.telefono or "",
                c.grupo_familiar or "",
                c.plan_ofrecido or "",
                c.estado or "",
                c.observaciones or "",
                c.conyuge or "",
                c.conyuge_edad or "",
                c.created_at.strftime("%d/%m/%Y") if c.created_at else "N/A"
            ]
            writer.writerow(row)
        
        output.seek(0)
        
        # Crear respuesta
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"mis_contactos_{timestamp}.csv"
        
        response = Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
        
        logging.info("Exportación CSV completada exitosamente")
        return response
        
    except Exception as e:
        logging.error(f"Error en exportación simple: {str(e)}")
        return jsonify({'error': 'Error en exportación simple'}), 500

@gestor.route('/exportar_mis_contactos')
@login_required
def exportar_mis_contactos():
    try:
        logging.info(f"Iniciando exportación para usuario {current_user.email}")
        
        # Forzar recarga de datos desde la base de datos
        db.session.expire_all()
        contactos = Contacto.query.filter_by(usuario_id=current_user.id).all()
        
        # Log para verificar cuántos contactos se están exportando
        logging.info(f"Exportando {len(contactos)} contactos para usuario {current_user.email}")
        
        # Agregar timestamp al nombre del archivo para evitar caché
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"mis_contactos_{timestamp}.xlsx"
        
        # Función para normalizar texto
        def normalize_text(text):
            if text:
                return text.strip().lower().capitalize()
            return text
        
        # Crear datos de forma más simple
        data = []
        for c in contactos:
            try:
                data.append({
                    "Origen": normalize_text(c.origen),
                    "Cobertura Actual": normalize_text(c.cobertura_actual_otra) if c.cobertura_actual == 'otros' else c.cobertura_actual,
                    "¿Por qué no toma la cobertura?": normalize_text(c.promocion) if c.promocion else "No especificado",
                    "Privado/Desregulado": c.privadoDesregulado,
                    "Apellido y nombre": c.apellido_nombre,
                    "Correo electrónico": c.correo_electronico,
                    "Edad titular": c.edad_titular,
                    "Teléfono": c.telefono,
                    "Grupo familiar": c.grupo_familiar,
                    "Plan ofrecido": c.plan_ofrecido,
                    "Estado": c.estado,
                    "Observaciones": c.observaciones,
                    "Cónyuge": c.conyuge,
                    "Edad cónyuge": c.conyuge_edad,
                    "Fecha de carga": c.created_at.strftime("%d/%m/%Y") if c.created_at else "N/A"
                })
            except Exception as e:
                logging.error(f"Error procesando contacto {c.id}: {str(e)}")
                continue
        
        logging.info(f"Datos procesados: {len(data)} registros")
        
        # Crear DataFrame de forma más segura
        try:
            df = pd.DataFrame(data)
            logging.info("DataFrame creado exitosamente")
        except Exception as e:
            logging.error(f"Error creando DataFrame: {str(e)}")
            return jsonify({'error': 'Error al procesar los datos'}), 500
        
        # Crear un nuevo libro de Excel
        try:
            wb = openpyxl.Workbook()
            logging.info("Libro Excel creado")
        except Exception as e:
            logging.error(f"Error creando libro Excel: {str(e)}")
            return jsonify({'error': 'Error al crear el archivo Excel'}), 500
        
        # Configurar estilos básicos
        try:
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            logging.info("Estilos configurados")
        except Exception as e:
            logging.error(f"Error configurando estilos: {str(e)}")
            # Continuar sin estilos si hay error
        
        # Crear la hoja de datos
        try:
            ws_data = wb.active
            ws_data.title = "Mis Contactos"
            
            # Agregar encabezados
            headers = list(df.columns)
            for col, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col)
                cell.value = header
                try:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                except:
                    pass  # Continuar sin estilos si hay error
            
            # Agregar datos
            for row_idx, row_data in enumerate(df.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_data.cell(row=row_idx, column=col_idx)
                    cell.value = str(value) if value is not None else ""
            
            logging.info("Datos agregados a Excel")
            
        except Exception as e:
            logging.error(f"Error agregando datos a Excel: {str(e)}")
            return jsonify({'error': 'Error al agregar datos al Excel'}), 500
        
        # Intentar guardar el archivo
        try:
            # Intentar guardar en el directorio actual primero
            wb.save(filename)
            logging.info(f"Archivo guardado como: {filename}")
            
            # Intentar enviar el archivo
            try:
                response = send_file(
                    filename, 
                    as_attachment=True, 
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                logging.info("Archivo enviado exitosamente")
                return response
            except Exception as e:
                logging.error(f"Error enviando archivo: {str(e)}")
                return jsonify({'error': 'Error al enviar el archivo'}), 500
                
        except Exception as e:
            logging.error(f"Error guardando archivo: {str(e)}")
            return jsonify({'error': 'Error al guardar el archivo Excel'}), 500
            
    except Exception as e:
        logging.error(f"Error general en exportación: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500

@gestor.route('/')
def redirigirlogin():
    return redirect(url_for('auth.login'))

@gestor.errorhandler(404)
def pageNotFound(error):
    return redirect(url_for('auth.login'))

@gestor.route('/verificar_email', methods=['POST'])
def verificar_email():
    email = request.json.get('email')
    tipo = request.json.get('tipo')  # 'login' o 'register'
    
    if not email:
        return jsonify({'valid': False, 'message': 'Email requerido'})
    
    # Verificar formato básico del email
    import re
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    if not email_pattern.match(email):
        return jsonify({'valid': False, 'message': 'Formato de email inválido'})
    
    # Verificar si existe en la base de datos
    usuario = Usuario.query.filter_by(email=email).first()
    
    if tipo == 'login':
        # Para login, el email debe existir
        if usuario:
            return jsonify({'valid': True, 'message': 'Email válido'})
        return jsonify({'valid': False, 'message': 'Email no registrado'})
    else:  # register
        # Para registro, el email no debe existir
        if usuario:
            return jsonify({'valid': False, 'message': 'Email ya registrado'})
        return jsonify({'valid': True, 'message': 'Email disponible'})

@gestor.route('/terminos-y-condiciones')
def terminos():
    return render_template('terminos.html')

@gestor.route('/robots.txt')
def robots_txt():
    """Servir archivo robots.txt para bloquear motores de búsqueda"""
    return Response(
        "User-agent: *\nDisallow: /\n\n"
        "User-agent: Googlebot\nDisallow: /\n\n"
        "User-agent: Bingbot\nDisallow: /\n\n"
        "User-agent: Slurp\nDisallow: /\n\n"
        "User-agent: Yandex\nDisallow: /\n\n"
        "User-agent: DuckDuckBot\nDisallow: /\n\n"
        "User-agent: Baiduspider\nDisallow: /\n\n"
        "User-agent: Sogou\nDisallow: /\n\n"
        "Noindex: /",
        mimetype='text/plain'
    )

@gestor.route('/sitemap.xml')
def sitemap_xml():
    """Servir sitemap.xml vacío para evitar indexación"""
    return Response(
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        '</urlset>',
        mimetype='application/xml'
    )

# Middleware para agregar headers de seguridad y privacidad
@gestor.before_request
def add_security_headers():
    """Agregar headers para evitar indexación y mejorar seguridad"""
    # Solo agregar headers si no es una solicitud de archivos estáticos o descargas
    if not request.path.startswith('/static/') and not request.path.startswith('/exportar'):
        response = make_response()
        response.headers['X-Robots-Tag'] = 'noindex, nofollow, noarchive, nosnippet, noimageindex'
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'DENY'
        response.headers['X-XSS-Protection'] = '1; mode=block'
        response.headers['Referrer-Policy'] = 'no-referrer'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'

# Middleware para manejar descargas en PythonAnywhere
@gestor.after_request
def add_download_headers(response):
    """Agregar headers específicos para descargas de archivos"""
    if request.path.startswith('/exportar'):
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        # Asegurar que el archivo se descargue correctamente
        if 'Content-Disposition' in response.headers:
            response.headers['Content-Disposition'] = response.headers['Content-Disposition'].replace('attachment; ', 'attachment; filename*=UTF-8\'\'')
    return response

if __name__ == '__main__':
    # Solo ejecutar en desarrollo local, no en PythonAnywhere
    import socket
    hostname = socket.gethostname()
    
    if 'liveconsole' not in hostname and 'pythonanywhere' not in hostname.lower():
        gestor.run(host='0.0.0.0', port=5555, debug=False)
    else:
        print("Aplicación configurada para PythonAnywhere")